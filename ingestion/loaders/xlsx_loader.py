"""工资表 Excel (.xlsx) 加载器模块。

将工资表 xlsx 的每行员工数据转成结构化自然语言段落，
以便向量召回能搜到类似"张三 3月工资"的查询。
"""

from __future__ import annotations

import os
import re

from langchain_core.documents import Document

from utils.logger_handler import logger


_HEADER_ALIASES: tuple[str, ...] = (
    "工号",
    "员工工号",
    "姓名",
    "员工姓名",
    "部门",
    "所属部门",
    "岗位",
    "基本工资",
    "岗位津贴",
    "绩效工资",
    "应发工资",
    "社保(个人)",
    "公积金(个人)",
    "个税",
    "个人所得税",
    "实发工资",
)


def _infer_month_from_filename(filename: str) -> tuple[int, int]:
    """从文件名提取 (year, month)。

    匹配示例：员工工资表_2026年1月.xlsx -> (2026, 1)

    Args:
        filename: 文件名（含扩展名或不含均可）。

    Returns:
        (year, month) 元组；没匹配到返回 (0, 0)。
    """
    match = re.search(r"(20\d{2})年(\d{1,2})月", filename)
    if match:
        return int(match.group(1)), int(match.group(2))
    return 0, 0


def _find_header_row(ws, *, max_scan_rows: int = 20) -> tuple[int, list[str]]:
    """在工作表前若干行中识别真正表头行。

    工资表通常前 1-2 行是标题/制表日期，不能直接把第一行当表头。
    """
    best_row_index = 0
    best_headers: list[str] = []
    best_score = 0

    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        headers = [str(value).strip() if value is not None else "" for value in row]
        normalized = {h.replace(" ", "").lower() for h in headers if h}
        score = 0
        for alias in _HEADER_ALIASES:
            if alias.replace(" ", "").lower() in normalized:
                score += 1
        if score > best_score:
            best_row_index = row_index
            best_headers = headers
            best_score = score

    if best_score >= 2:
        return best_row_index, best_headers
    return 0, []


def _guess_doc_department(filename: str, sheet_name: str) -> str:
    """从文件名/sheet名猜部门。

    规则：
    - 包含"销售" → "销售部"
    - 包含"技术/研发" → "研发部"
    - 包含"财务" → "财务部"
    - 包含"HR/人事" → "人力资源部"
    - 包含"行政" → "行政部"
    - 否则返回 ""

    Args:
        filename: 文件名。
        sheet_name: sheet 名称。

    Returns:
        推测的部门名称，推测不到返回空字符串。
    """
    combined = f"{filename} {sheet_name}"
    if "销售" in combined:
        return "销售部"
    if "技术" in combined or "研发" in combined:
        return "研发部"
    if "财务" in combined:
        return "财务部"
    if "HR" in combined.upper() or "人事" in combined:
        return "人力资源部"
    if "行政" in combined:
        return "行政部"
    return ""


def _row_to_nl(
    headers: list[str],
    row_values: list,
    *,
    year: int,
    month: int,
    department: str,
    source_file: str,
    sheet: str,
) -> tuple[str, dict]:
    """把一行工资数据转成自然语言段 + 业务 metadata。

    Args:
        headers: 表头列表。
        row_values: 行数据列表，与 headers 一一对应。
        year: 数据年份（从文件名推断，可能为0）。
        month: 数据月份（从文件名推断，可能为0）。
        department: 部门名称（推测的，可能为空）。
        source_file: 源文件路径。
        sheet: sheet 名称。

    Returns:
        (自然语言段落文本, metadata 字典)
    """
    name_map = {
        ("姓名", "员工姓名"): "employee_name",
        ("工号", "员工工号"): "employee_id",
        ("部门", "所属部门"): "employee_department",
        ("应发合计", "应发工资", "总工资"): "gross_pay",
        ("实发", "实发工资", "到手", "到手工资"): "net_pay",
        ("个税", "个人所得税"): "personal_income_tax",
    }

    def _find_value(targets: tuple[str, ...], h_list: list[str], v_list: list) -> str | None:
        for h, v in zip(h_list, v_list):
            h_lower = str(h).strip().lower()
            for t in targets:
                if t.lower() == h_lower:
                    return str(v).strip() if v is not None else ""
        return None

    header_lower = [str(h).strip().lower() for h in headers]
    meta: dict = {}
    for keys, meta_key in name_map.items():
        found = _find_value(keys, headers, row_values)
        if found is not None and found != "":
            meta[meta_key] = found

    display_name = meta.get("employee_name", "未知员工")
    ym_title = f"【{year}年{month}月工资 - {display_name}】" if year and month else f"【工资 - {display_name}】"

    nl_lines: list[str] = [ym_title]
    dept_display = department or meta.get("employee_department", "") or ""
    for h, v in zip(headers, row_values):
        val = str(v).strip() if v is not None else ""
        if val == "":
            continue
        h_str = str(h).strip()
        if h_str == "":
            continue
        suffix = ""
        if any(k in h_str for k in ["工资", "补贴", "合计", "社保", "公积金", "个税", "税", "补", "金"]):
            suffix = " 元"
        elif "天" in h_str or "天数" in h_str:
            suffix = " 天"
        nl_lines.append(f"{h_str}：{val}{suffix}")

    if dept_display and "部门：" not in "\n".join(nl_lines):
        nl_lines.insert(1, f"部门：{dept_display}")

    page_content = "\n".join(nl_lines)

    doc_dept = department or "全体"
    meta_dict: dict = {
        "doc_type": "salary",
        "sensitivity_level": "confidential",
        "department": doc_dept,
        "owner_role": "self",
        "source_file": source_file,
        "sheet_name": sheet,
        "data_year": year,
        "data_month": month,
    }
    meta_dict.update(meta)

    return page_content, meta_dict


def xlsx_loader(filepath: str) -> list[Document]:
    """加载工资表 xlsx，返回多个 Document（每行一个）。

    使用 openpyxl 的只读模式，避免大文件内存占用。
    缺依赖时提示 pip install openpyxl 并返回 []。

    Args:
        filepath: Excel 文件路径。

    Returns:
        每行工资数据对应的 Document 列表；失败或无数据返回 []。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("[xlsx_loader] 缺少 openpyxl 依赖，请执行: pip install openpyxl")
        return []

    try:
        filename = os.path.basename(filepath)
        year, month = _infer_month_from_filename(filename)

        wb = load_workbook(filepath, read_only=True, data_only=True)
        result_docs: list[Document] = []

        try:
            for ws in wb.worksheets:
                sheet_name = ws.title
                dept = _guess_doc_department(filename, sheet_name)

                header_row_index, headers = _find_header_row(ws)
                if header_row_index <= 0 or not headers:
                    logger.warning(
                        f"[xlsx_loader] {filepath} sheet={sheet_name} 未识别到工资表头，跳过"
                    )
                    continue

                for row in ws.iter_rows(min_row=header_row_index + 1):
                    row_values = [
                        str(cell.value).strip() if cell.value is not None else "" for cell in row
                    ]
                    if all(v == "" for v in row_values):
                        continue

                    page_content, meta_dict = _row_to_nl(
                        headers,
                        row_values,
                        year=year,
                        month=month,
                        department=dept,
                        source_file=filepath,
                        sheet=sheet_name,
                    )
                    result_docs.append(Document(page_content=page_content, metadata=meta_dict))
        finally:
            wb.close()

        logger.info(f"[xlsx_loader] {filepath} 提取到 {len(result_docs)} 条工资行文档")

        if not result_docs:
            logger.warning(f"[xlsx_loader] xlsx 没提取到行: {filepath}")

        return result_docs

    except Exception as e:
        logger.exception(f"[xlsx_loader] 加载 xlsx 异常: {e}")
        return []
